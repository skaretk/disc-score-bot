import openpyxl
import os
from score.playeralias import PlayerAlias

class ExcelFile:
    def __init__(self):
        self.sheet_name = 'MemeberArchive'
        self.first_name = 'Fornavn'
        self.first_name_column = ''
        self.surname = 'Etternavn'
        self.surname_column = ''
        self.contingent = 'Kontingent'
        self.contingent_column = ''

class Members:
    def __init__(self, path, file):
        self.path = f'{os.getcwd()}/cfg/{path}'
        self.file = file
        self.is_member = 'Betalt'
        self.member_list = []
        self.excel_file = ExcelFile()

    def parse(self):
        wb = openpyxl.load_workbook(filename=f'{self.path}/{self.file}', read_only=True)
        sheet = wb.active
        print(sheet.title)

        # Fetch header
        for cell in sheet[1]:
            if cell.value == self.excel_file.first_name:
                self.excel_file.first_name_column = cell.column
            elif cell.value == self.excel_file.surname:
                self.excel_file.surname_column = cell.column
            elif cell.value == self.excel_file.contingent:
                self.excel_file.contingent_column = cell.column

        # Fetch all members
        for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
            first_name = ''
            surname = ''
            member = False
            for cell in row:
                if cell.column == self.excel_file.first_name_column:
                    first_name = cell.value
                elif cell.column == self.excel_file.surname_column:
                    surname = cell.value
                elif cell.column == self.excel_file.contingent_column:
                    if cell.value == self.is_member:
                        member = True
            # Paid ?
            if member is True:
                self.member_list.append(PlayerAlias(f'{first_name} {surname}'))
        wb.close()
